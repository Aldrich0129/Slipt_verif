# -*- coding: utf-8 -*-
"""
PDF Payslip Validator - 核心验证模块
不依赖GUI，纯逻辑处理
"""

import os
import re
import unicodedata
from datetime import datetime
import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ===== 工具函数 =====
def strip_accents(s: str) -> str:
    """去除重音符号"""
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')


def normalize_name(name: str) -> str:
    """标准化姓名：去重音、转大写、去多余空格"""
    name = strip_accents(name)
    name = name.upper()
    name = re.sub(r'\s+', ' ', name)
    name = name.strip()
    # 去除逗号和下划线
    name = name.replace(',', ' ').replace('_', ' ')
    name = re.sub(r'\s+', ' ', name)
    return name.strip()


def parse_filename(filename: str) -> dict:
    """
    解析文件名
    格式: 809_MARTINEZ MONTERO_ LAURA MARIA_Payslip_01092024.pdf
    返回: {'codigo': '809', 'nombre': 'MARTINEZ MONTERO LAURA MARIA', 'fecha': '01092024'}
    """
    # 移除.pdf扩展名
    name = filename.replace('.pdf', '').replace('.PDF', '')

    # 尝试匹配格式: 编号_姓名_Payslip_日期
    # 编号可能是2-6位数字
    pattern = r'^(\d{2,6})_(.*?)_Payslip_(.+)$'
    match = re.match(pattern, name, re.IGNORECASE)

    if match:
        codigo = match.group(1)
        nombre = normalize_name(match.group(2))
        fecha = match.group(3).replace('_', '')  # 移除可能的下划线

        return {
            'codigo': codigo,
            'nombre': nombre,
            'fecha': fecha,
            'valid': True
        }
    else:
        return {
            'codigo': '',
            'nombre': '',
            'fecha': '',
            'valid': False,
            'error': '文件名格式不符合规范'
        }


def extract_pdf_info(pdf_path: str) -> dict:
    """
    从PDF中提取关键信息
    返回: {'codigo': '809', 'nombre': 'MARTINEZ MONTERO, LAURA MARIA', 'nif': '...', 'periodo': '...'}
    """
    try:
        with fitz.open(pdf_path) as doc:
            # 通常信息在第一页
            if doc.page_count > 0:
                page = doc[0]
                text = page.get_text('text')

                # 提取编号（809/1格式）
                codigo = None
                codigo_pattern = r'\n\s*(\d{2,6})\/\d\s*\n'
                match = re.search(codigo_pattern, text)
                if match:
                    codigo = match.group(1)

                # 如果没找到，尝试其他模式
                if not codigo:
                    # 尝试查找单独的编号行
                    for line in text.split('\n')[:30]:
                        line = line.strip()
                        if re.match(r'^\d{2,6}$', line):
                            codigo = line
                            break

                # 提取姓名（通常是大写的完整姓名）
                nombre = None
                nombre_pattern = r'\n\s*\d{2,6}\/\d\s*\n\s*([A-ZÁÉÍÓÚÜÑ ,.\'\\-]+?)\s*\n'
                match = re.search(nombre_pattern, text)
                if match:
                    nombre = normalize_name(match.group(1))

                # 如果没找到，查找包含逗号的大写行（姓名通常包含逗号）
                if not nombre:
                    lines = text.split('\n')[:60]
                    for line in lines:
                        line = line.strip()
                        if ',' in line and line.isupper() and len(line.split()) >= 2:
                            # 确保不是地址或其他信息
                            if 'CL ' not in line and 'PZ ' not in line and 'BARCELONA' not in line:
                                nombre = normalize_name(line)
                                break

                # 提取NIF
                nif = None
                nif_pattern = r'N\.I\.F\.\s*([A-Z0-9]{8,})'
                match = re.search(nif_pattern, text)
                if match:
                    nif = match.group(1)

                # 提取期间
                periodo = None
                periodo_pattern = r'PERÍODO\s+(\d{1,2}\s+\w+\s+\d{1,2}\s+\w+\s+\d{4})'
                match = re.search(periodo_pattern, text, re.IGNORECASE)
                if match:
                    periodo = match.group(1)

                # 提取Nº. Afiliación S.S.
                afiliacion = None
                afiliacion_pattern = r'Afiliación\s+S\.S\.\s+(\d+)'
                match = re.search(afiliacion_pattern, text, re.IGNORECASE)
                if match:
                    afiliacion = match.group(1)

                return {
                    'codigo': codigo or 'NO ENCONTRADO',
                    'nombre': nombre or 'NO ENCONTRADO',
                    'nif': nif or 'NO ENCONTRADO',
                    'periodo': periodo or 'NO ENCONTRADO',
                    'afiliacion': afiliacion or 'NO ENCONTRADO',
                    'valid': True
                }
            else:
                return {
                    'valid': False,
                    'error': 'PDF vacío'
                }

    except Exception as e:
        return {
            'valid': False,
            'error': f'Error al leer PDF: {str(e)}'
        }


def compare_info(filename_info: dict, pdf_info: dict) -> dict:
    """
    比较文件名和PDF内容
    返回验证结果
    """
    result = {
        'codigo_match': False,
        'nombre_match': False,
        'overall_match': False,
        'errors': []
    }

    # 比较编号
    if filename_info.get('codigo') and pdf_info.get('codigo'):
        if filename_info['codigo'] == pdf_info['codigo']:
            result['codigo_match'] = True
        else:
            result['errors'].append(f"编号不匹配: 文件名={filename_info['codigo']}, PDF={pdf_info['codigo']}")
    else:
        result['errors'].append("编号信息缺失")

    # 比较姓名（标准化后比较）
    if filename_info.get('nombre') and pdf_info.get('nombre'):
        fn_nombre = normalize_name(filename_info['nombre'])
        pdf_nombre = normalize_name(pdf_info['nombre'])

        # 完全匹配或高度相似
        if fn_nombre == pdf_nombre:
            result['nombre_match'] = True
        else:
            # 检查是否包含关系（因为格式可能略有不同）
            fn_parts = set(fn_nombre.split())
            pdf_parts = set(pdf_nombre.split())

            # 如果至少80%的词匹配，认为相似
            if fn_parts and pdf_parts:
                intersection = fn_parts & pdf_parts
                union = fn_parts | pdf_parts
                similarity = len(intersection) / len(union)

                if similarity >= 0.8:
                    result['nombre_match'] = True
                    result['errors'].append(f"姓名部分匹配(相似度{similarity:.0%}): 文件名={fn_nombre}, PDF={pdf_nombre}")
                else:
                    result['errors'].append(f"姓名不匹配: 文件名={fn_nombre}, PDF={pdf_nombre}")
            else:
                result['errors'].append(f"姓名不匹配: 文件名={fn_nombre}, PDF={pdf_nombre}")
    else:
        result['errors'].append("姓名信息缺失")

    # 总体匹配：编号和姓名都匹配
    result['overall_match'] = result['codigo_match'] and result['nombre_match']

    return result


def validate_folder(folder_path: str, progress_callback=None) -> list:
    """
    验证文件夹中的所有PDF文件
    返回验证结果列表
    """
    results = []

    # 获取所有PDF文件
    pdf_files = [f for f in os.listdir(folder_path) if f.lower().endswith('.pdf')]
    total_files = len(pdf_files)

    if total_files == 0:
        return results

    for idx, filename in enumerate(pdf_files):
        if progress_callback:
            progress_callback(idx + 1, total_files, filename)

        pdf_path = os.path.join(folder_path, filename)

        # 解析文件名
        filename_info = parse_filename(filename)

        # 提取PDF信息
        pdf_info = extract_pdf_info(pdf_path)

        # 比较信息
        if filename_info.get('valid') and pdf_info.get('valid'):
            comparison = compare_info(filename_info, pdf_info)
        else:
            comparison = {
                'codigo_match': False,
                'nombre_match': False,
                'overall_match': False,
                'errors': []
            }
            if not filename_info.get('valid'):
                comparison['errors'].append(filename_info.get('error', '文件名解析失败'))
            if not pdf_info.get('valid'):
                comparison['errors'].append(pdf_info.get('error', 'PDF解析失败'))

        # 汇总结果
        result = {
            'filename': filename,
            'fn_codigo': filename_info.get('codigo', ''),
            'fn_nombre': filename_info.get('nombre', ''),
            'fn_fecha': filename_info.get('fecha', ''),
            'pdf_codigo': pdf_info.get('codigo', ''),
            'pdf_nombre': pdf_info.get('nombre', ''),
            'pdf_nif': pdf_info.get('nif', ''),
            'pdf_periodo': pdf_info.get('periodo', ''),
            'pdf_afiliacion': pdf_info.get('afiliacion', ''),
            'codigo_match': comparison['codigo_match'],
            'nombre_match': comparison['nombre_match'],
            'overall_match': comparison['overall_match'],
            'errors': '; '.join(comparison['errors']) if comparison['errors'] else ''
        }

        results.append(result)

    return results


def generate_excel_report(results: list, output_path: str):
    """
    生成Excel报告，用颜色标记验证结果
    绿色：完全匹配
    红色：不匹配
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "验证报告"

    # 定义样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 表头
    headers = [
        '文件名',
        '文件名-编号',
        '文件名-姓名',
        '文件名-日期',
        'PDF-编号',
        'PDF-姓名',
        'PDF-NIF',
        'PDF-期间',
        'PDF-社保号',
        '编号匹配',
        '姓名匹配',
        '验证结果',
        '错误说明'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # 数据行
    for row_num, result in enumerate(results, 2):
        # 基本信息
        ws.cell(row=row_num, column=1, value=result['filename']).border = border
        ws.cell(row=row_num, column=2, value=result['fn_codigo']).border = border
        ws.cell(row=row_num, column=3, value=result['fn_nombre']).border = border
        ws.cell(row=row_num, column=4, value=result['fn_fecha']).border = border
        ws.cell(row=row_num, column=5, value=result['pdf_codigo']).border = border
        ws.cell(row=row_num, column=6, value=result['pdf_nombre']).border = border
        ws.cell(row=row_num, column=7, value=result['pdf_nif']).border = border
        ws.cell(row=row_num, column=8, value=result['pdf_periodo']).border = border
        ws.cell(row=row_num, column=9, value=result['pdf_afiliacion']).border = border

        # 编号匹配
        codigo_cell = ws.cell(row=row_num, column=10, value='✓' if result['codigo_match'] else '✗')
        codigo_cell.border = border
        codigo_cell.alignment = Alignment(horizontal='center')
        codigo_cell.fill = green_fill if result['codigo_match'] else red_fill

        # 姓名匹配
        nombre_cell = ws.cell(row=row_num, column=11, value='✓' if result['nombre_match'] else '✗')
        nombre_cell.border = border
        nombre_cell.alignment = Alignment(horizontal='center')
        nombre_cell.fill = green_fill if result['nombre_match'] else red_fill

        # 总体结果
        overall_cell = ws.cell(row=row_num, column=12, value='匹配' if result['overall_match'] else '不匹配')
        overall_cell.border = border
        overall_cell.alignment = Alignment(horizontal='center')
        overall_cell.fill = green_fill if result['overall_match'] else red_fill
        overall_cell.font = Font(bold=True)

        # 错误说明
        error_cell = ws.cell(row=row_num, column=13, value=result['errors'])
        error_cell.border = border
        error_cell.alignment = Alignment(wrap_text=True)
        if result['errors']:
            error_cell.fill = yellow_fill

    # 调整列宽
    column_widths = [35, 12, 30, 15, 12, 30, 15, 25, 18, 10, 10, 12, 50]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 添加统计信息
    total = len(results)
    matched = sum(1 for r in results if r['overall_match'])
    unmatched = total - matched

    stats_row = len(results) + 3
    ws.cell(row=stats_row, column=1, value="统计信息").font = Font(bold=True, size=12)
    ws.cell(row=stats_row + 1, column=1, value=f"总文件数: {total}")
    ws.cell(row=stats_row + 2, column=1, value=f"匹配: {matched}").fill = green_fill
    ws.cell(row=stats_row + 3, column=1, value=f"不匹配: {unmatched}").fill = red_fill
    ws.cell(row=stats_row + 4, column=1, value=f"匹配率: {matched/total*100:.1f}%" if total > 0 else "N/A")

    # 保存
    wb.save(output_path)
